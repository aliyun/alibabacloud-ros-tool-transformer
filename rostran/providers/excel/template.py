import logging
from typing import List

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from rostran.core.exceptions import TemplateFormatNotSupport, InvalidExcelTemplate
from rostran.core.format import FileFormat
from rostran.core.parameters import Parameters, Parameter
from rostran.core.resources import Resources, Resource
from rostran.core.properties import Property
from rostran.core.template import Template, RosTemplate

logger = logging.getLogger(__name__)

MAX_TEMPLATES = 5


class ExcelTemplate(Template):
    SECTIONS = (PARAMETERS, RESOURCES) = (
        "ROS::Parameters",
        "ROS::Resources",
    )

    @classmethod
    def initialize(cls, path: str, format: FileFormat = FileFormat.Excel):
        if format != FileFormat.Excel:
            raise TemplateFormatNotSupport(path=path, format=format)

        source = load_workbook(path, read_only=True, data_only=True)
        return cls(source=source)

    def transform(self) -> List[RosTemplate]:
        logger.info(f"Transform excel template to ROS template")

        # check column
        sheet: Worksheet = self.source.active
        max_column = sheet.max_column
        if max_column > MAX_TEMPLATES + 1 or max_column < 2:
            reason = (
                f"Column number {max_column} not meet limit(>=2,<={MAX_TEMPLATES + 1}"
            )
            raise InvalidExcelTemplate(reason=reason)

        return self._transform()

    def _transform(self) -> List[RosTemplate]:
        sheet: Worksheet = self.source.active
        max_column = sheet.max_column

        templates = []
        for _ in range(max_column - 1):
            templates.append(RosTemplate())

        cur_section = None
        cur_resources = None
        section_occurs = {}
        for row in sheet.rows:
            header_cell: Cell = row[0]
            header_value = header_cell.value
            if isinstance(header_value, str):
                header_value = header_value.strip()
            header_value_str = str(header_value)

            # Section
            if header_value_str.startswith("#"):
                continue

            if header_value_str.startswith("ROS::"):
                if header_value_str not in self.SECTIONS:
                    reason = f"Section {header_value} on {header_cell} is invalid. Allowed sections: {self.SECTIONS}"
                    raise InvalidExcelTemplate(reason=reason)
                cur_section = header_value_str
                continue

            if cur_section and not header_value:
                cur_section = None

            if not cur_section:
                continue

            # Parameters
            if cur_section == self.PARAMETERS:
                if (
                    section_occurs.get(header_value_str)
                    and header_value_str == cur_section
                ):
                    raise InvalidExcelTemplate(
                        reason=f"Section {header_value_str} on {header_cell} is duplicated"
                    )
                else:
                    section_occurs[header_value_str] = True
                for i, cell in enumerate(row[1:max_column]):
                    parameter = Parameter.initialize_from_excel(header_cell, cell)
                    templates[i].parameters.add(parameter)

            # Resources
            if cur_section == self.RESOURCES:
                if (
                    section_occurs.get(header_value_str)
                    and header_value_str == cur_section
                ):
                    raise InvalidExcelTemplate(
                        reason=f"Section {header_value_str} on {header_cell} is duplicated"
                    )
                else:
                    section_occurs[header_value_str] = True

                # Specific resource
                if "::" in header_value_str.split("\n")[0]:
                    cur_resources = []
                    for i, cell in enumerate(row[1:max_column]):
                        cell_value = (
                            "" if cell.value is None else str(cell.value).strip()
                        )
                        if cell_value:
                            resource = Resource.initialize_from_excel(header_cell, cell)
                            templates[i].resources.add(resource)
                        else:
                            resource = None
                        cur_resources.append(resource)
                else:
                    for i, cell in enumerate(row[1:max_column]):
                        resource = cur_resources[i]
                        if not resource:
                            continue
                        prop = Property.initialize_from_excel(header_cell, cell)
                        resource.properties.add(prop)

        for template in templates:
            for res in template.resources.values():
                res: Resource
                res.properties = res.properties.resolve()
        return templates

    def _handle_parameters(self, parameters: Parameters, cell: Cell):
        pass

    def _handle_resources(self, resources: Resources, cell: Cell):
        pass
